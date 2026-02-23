import pandas as pd
import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import OneHotEncoder, StandardScaler
from sklearn.compose import ColumnTransformer
from sklearn.pipeline import Pipeline
from sklearn.linear_model import LogisticRegression
from sklearn.tree import DecisionTreeClassifier
from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier
from sklearn.svm import SVC
from sklearn.neighbors import KNeighborsClassifier
from sklearn.metrics import (accuracy_score, precision_score, recall_score, f1_score,
                             confusion_matrix, roc_curve, auc)
import seaborn as sns
import matplotlib.pyplot as plt
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

df_clean = pd.read_excel("sprint 2/AI_Electra_Cleaned_dataset.xlsx")
print("Step 1: Dataset loaded successfully")

df_clean["Cancer"] = (df_clean["Tumor_Marker_Level"] > 35).astype(int)
print("Step 2: Target variable 'Cancer' created")

drop_cols = ["Patient_ID", "First_Name", "Last_Name",
             "Doctor_Assigned", "Admission_Date", "Discharge_Date",
             "Hospital_Department", "Insurance_Provider",
             "Tumor_Marker_Level"]
df = df_clean.drop(columns=drop_cols)
print("Step 3: Irrelevant columns dropped")

df[["Systolic_BP", "Diastolic_BP"]] = df["Blood_Pressure"].str.split("/", expand=True).astype(float)
df.drop(columns=["Blood_Pressure"], inplace=True)
df.drop(columns=["Height_cm", "Weight_kg"], inplace=True)

def bmi_category(bmi):
    if bmi < 18.5: return "Underweight"
    elif 18.5 <= bmi < 25: return "Normal"
    elif 25 <= bmi < 30: return "Overweight"
    else: return "Obese"
df["BMI_Category"] = df["BMI"].apply(bmi_category)
df["Hypertension_Flag"] = np.where((df["Systolic_BP"] >= 140) | (df["Diastolic_BP"] >= 90), "Yes", "No")

print("Step 4: Feature engineering completed")

X = df.drop("Cancer", axis=1)
y = df["Cancer"]

categorical_cols = X.select_dtypes(include=["object"]).columns.tolist()
numeric_cols = X.select_dtypes(include=["int64", "float64"]).columns.tolist()

print(f"Step 5: Features separated into {len(numeric_cols)} numeric and {len(categorical_cols)} categorical")
print("Class Distribution:\n", y.value_counts())

preprocessor = ColumnTransformer(
    transformers=[
        ("num", StandardScaler(), numeric_cols),
        ("cat", OneHotEncoder(handle_unknown="ignore"), categorical_cols)
    ]
)

pipeline = Pipeline(steps=[("preprocessor", preprocessor)])
X_processed = pipeline.fit_transform(X)

cat_features = pipeline.named_steps["preprocessor"].named_transformers_["cat"].get_feature_names_out(categorical_cols)
all_features = numeric_cols + cat_features.tolist()

X_processed_df = pd.DataFrame(
    X_processed.toarray() if hasattr(X_processed, "toarray") else X_processed,
    columns=all_features
)
print("Step 6: Preprocessing applied (scaling + encoding)")

X_train, X_test, y_train, y_test = train_test_split(
    X_processed_df, y, test_size=0.2, random_state=42, stratify=y
)
print("Step 7: Train-test split completed (80/20)")

models = {
    "Logistic Regression": LogisticRegression(max_iter=1000, class_weight="balanced"),
    "Decision Tree": DecisionTreeClassifier(random_state=42, class_weight="balanced"),
    "Random Forest": RandomForestClassifier(n_estimators=100, random_state=42, class_weight="balanced"),
    "SVM": SVC(kernel="rbf", probability=True, random_state=42, class_weight="balanced"),
    "KNN": KNeighborsClassifier(n_neighbors=5),
    "Gradient Boosting": GradientBoostingClassifier(random_state=42)
}

metrics = []
roc_data = {}
os.makedirs("ConfusionMatrices", exist_ok=True)

for name, model in models.items():
    model.fit(X_train, y_train)
    y_pred = model.predict(X_test)
    y_proba = model.predict_proba(X_test)[:, 1] if hasattr(model, "predict_proba") else None

    acc = accuracy_score(y_test, y_pred)
    prec = precision_score(y_test, y_pred)
    rec = recall_score(y_test, y_pred)
    f1 = f1_score(y_test, y_pred)
    cm = confusion_matrix(y_test, y_pred)
    tn, fp, fn, tp = cm.ravel()

    metrics.append([name, acc, prec, rec, f1, tn, fp, fn, tp])

    if y_proba is not None:
        fpr, tpr, _ = roc_curve(y_test, y_proba)
        roc_auc = auc(fpr, tpr)
        roc_data[name] = (fpr, tpr, roc_auc)

    plt.figure(figsize=(4, 3))
    sns.heatmap(cm, annot=True, fmt="d", cmap="Blues",
                xticklabels=["No Cancer", "Cancer"],
                yticklabels=["No Cancer", "Cancer"])
    plt.title(f"Confusion Matrix - {name}")
    plt.xlabel("Predicted")
    plt.ylabel("Actual")
    plt.savefig(f"ConfusionMatrices/ConfusionMatrix_{name.replace(' ', '_')}.png")
    plt.close()

    print(f"Step 8: Model {name} trained and evaluated")

with pd.ExcelWriter("sprint 2(new)/AI_Electra_ProcessedDataset.xlsx") as writer:
    train_data = X_train.copy()
    train_data["Cancer"] = y_train.values
    train_data.to_excel(writer, sheet_name="Train", index=False)

    test_data = X_test.copy()
    test_data["Cancer"] = y_test.values
    test_data.to_excel(writer, sheet_name="Test", index=False)

    processed_dataset = X_processed_df.copy()
    processed_dataset["Cancer"] = y.values
    processed_dataset.to_excel(writer, sheet_name="Processed_Dataset", index=False)

print("Step 9: Train, test, and processed dataset saved to Excel")

df_metrics = pd.DataFrame(metrics, columns=[
    "Model", "Accuracy", "Precision", "Recall", "F1",
    "TN (True Negative)", "FP (False Positive)",
    "FN (False Negative)", "TP (True Positive)"
])

avg_row = ["Average"] + df_metrics.iloc[:, 1:].mean().tolist()
df_metrics.loc[len(df_metrics)] = avg_row

df_metrics.to_excel("sprint 2(new)/AI_Electra_ModelMetrics.xlsx", index=False)

wb = load_workbook("sprint 2(new)/AI_Electra_ModelMetrics.xlsx")
ws = wb.active
row_offset = len(df_metrics) + 3

for i, name in enumerate(models.keys(), start=0):
    img_path = f"ConfusionMatrices/ConfusionMatrix_{name.replace(' ', '_')}.png"
    img = XLImage(img_path)
    img.width, img.height = 300, 220
    ws.add_image(img, f"A{row_offset + i*12}")
    ws[f"A{row_offset + i*12 - 1}"] = f"{name} Confusion Matrix"

df_melted = df_metrics.melt(
    id_vars=["Model"],
    value_vars=["Accuracy", "Precision", "Recall", "F1"],
    var_name="Metric",
    value_name="Score"
)

plt.figure(figsize=(8, 6))
sns.lineplot(data=df_melted, x="Model", y="Score", hue="Metric", marker="o")
plt.ylim(0, 1)
plt.title("Model Performance Comparison (Line Graph)")
plt.savefig("sprint 2(new)/Model_Performance_Line.png")
plt.close()

plt.figure(figsize=(8, 6))
sns.barplot(data=df_melted, x="Model", y="Score", hue="Metric")
plt.ylim(0, 1)
plt.title("Model Performance Comparison (Bar Graph)")
plt.savefig("sprint 2(new)/Model_Performance_Bar.png")
plt.close()

final_row = ws.max_row + 3
line_img = XLImage("sprint 2(new)/Model_Performance_Line.png")
line_img.width, line_img.height = 480, 320
ws.add_image(line_img, f"A{final_row}")
ws[f"A{final_row - 1}"] = "Model Performance Line Graph"

bar_img = XLImage("sprint 2(new)/Model_Performance_Bar.png")
bar_img.width, bar_img.height = 480, 320
ws.add_image(bar_img, f"A{final_row + 20}")
ws[f"A{final_row + 19}"] = "Model Performance Bar Graph"

rf_model = models["Random Forest"]
importances = rf_model.feature_importances_
feat_imp = pd.Series(importances, index=X_train.columns).sort_values(ascending=False)

plt.figure(figsize=(8, 6))
sns.barplot(x=feat_imp.values[:15], y=feat_imp.index[:15])
plt.title("Top 15 Feature Importances (Random Forest)")
plt.savefig("sprint 2(new)/Feature_Importance.png")
plt.close()

feat_img = XLImage("sprint 2(new)/Feature_Importance.png")
feat_img.width, feat_img.height = 480, 320
ws.add_image(feat_img, f"A{final_row + 40}")
ws[f"A{final_row + 39}"] = "Feature Importance (Random Forest)"

plt.figure(figsize=(8, 6))
for name, (fpr, tpr, roc_auc) in roc_data.items():
    plt.plot(fpr, tpr, label=f"{name} (AUC={roc_auc:.2f})")
plt.plot([0, 1], [0, 1], linestyle="--", color="gray")
plt.xlabel("False Positive Rate")
plt.ylabel("True Positive Rate (Recall)")
plt.title("ROC Curves")
plt.legend()
plt.savefig("sprint 2(new)/ROC_Curves.png")
plt.close()

roc_img = XLImage("sprint 2(new)/ROC_Curves.png")
roc_img.width, roc_img.height = 480, 320
ws.add_image(roc_img, f"A{final_row + 60}")
ws[f"A{final_row + 59}"] = "ROC Curves with AUC"

summary = wb.create_sheet("Summary")
summary["A1"] = "Summary of Data & Feature Engineering"
summary["A3"] = f"Total Records: {len(df)}"
summary["A4"] = f"Train Size: {len(X_train)}, Test Size: {len(X_test)}"
summary["A5"] = "Class Distribution (0=No Cancer, 1=Cancer):"
summary["A6"] = str(y.value_counts().to_dict())
summary["A8"] = "Dropped Columns: Patient_ID, First_Name, Last_Name, Doctor_Assigned, Admission_Date, Discharge_Date, Hospital_Department, Insurance_Provider, Tumor_Marker_Level"
summary["A10"] = "Engineered Features: Systolic_BP, Diastolic_BP, BMI_Category, Hypertension_Flag"
summary["A12"] = "Note: Tumor Marker was excluded to avoid data leakage."

wb.save("sprint 2(new)/AI_Electra_ModelMetrics.xlsx")
print("Step 14: Summary sheet added and Excel saved")

print("Pipeline completed successfully!")
print("Deliverables generated:")
print(" - AI_Electra_ProcessedDataset.xlsx (Train, Test, Processed Dataset)")
print(" - AI_Electra_ModelMetrics.xlsx (Metrics, Confusion Matrices, Graphs, Feature Importance, ROC, Summary)")
print(" - PNG files in 'sprint 2' + ConfusionMatrices folder")
