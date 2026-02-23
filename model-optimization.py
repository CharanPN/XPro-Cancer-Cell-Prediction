# ================================================
# Week 5 - Optimized Model with PCA, RandomizedSearchCV, Metrics, Plots, and Excel Deliverable
# ================================================
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.model_selection import RandomizedSearchCV
from sklearn.linear_model import LogisticRegression
from sklearn.tree import DecisionTreeClassifier
from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier
from sklearn.svm import SVC
from sklearn.neighbors import KNeighborsClassifier
from sklearn.metrics import (
    accuracy_score, precision_score, recall_score, f1_score,
    confusion_matrix, ConfusionMatrixDisplay, roc_curve, auc
)
from sklearn.decomposition import PCA
from sklearn.pipeline import Pipeline
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage
import os
import numpy as np

# -----------------------
# Step 0: Create folders for plots
# -----------------------
os.makedirs("sprint3/Week5_Plots", exist_ok=True)

# -----------------------
# Step 1: Load Train/Test Data
# -----------------------
train_df = pd.read_excel("sprint3/AI_Electra_ProcessedDataset.xlsx", sheet_name="Train")
test_df = pd.read_excel("sprint3/AI_Electra_ProcessedDataset.xlsx", sheet_name="Test")

X_train = train_df.drop("Cancer", axis=1)
y_train = train_df["Cancer"]
X_test = test_df.drop("Cancer", axis=1)
y_test = test_df["Cancer"]

print("Train/Test datasets loaded successfully...")

# -----------------------
# Step 2: Define Pipelines & Hyperparameter Grids with PCA
# -----------------------
param_grids = {
    "Logistic Regression": {
        "pipeline": Pipeline([
            ("pca", PCA(random_state=42)),
            ("model", LogisticRegression(class_weight="balanced", max_iter=1000, random_state=42))
        ]),
        "params": {
            "pca__n_components": [0.80, 0.85, 0.90, 0.95, 0.99],
            "model__C": np.linspace(0.01, 1, 20),
            "model__solver": ["liblinear", "lbfgs"]
        }
    },
    "Decision Tree": {
        "pipeline": Pipeline([
            ("pca", PCA(random_state=42)),
            ("model", DecisionTreeClassifier(class_weight="balanced", random_state=42))
        ]),
        "params": {
            "pca__n_components": [0.80, 0.85, 0.90, 0.95, 0.99],
            "model__max_depth": [None, 5, 10, 20],
            "model__min_samples_split": [2, 5, 10],
            "model__min_samples_leaf": [1, 2, 4]
        }
    },
    "Random Forest": {
        "pipeline": Pipeline([
            ("pca", PCA(random_state=42)),
            ("model", RandomForestClassifier(class_weight="balanced", random_state=42))
        ]),
        "params": {
            "pca__n_components": [0.80, 0.85, 0.90, 0.95, 0.99],
            "model__n_estimators": [100, 200, 300],
            "model__max_depth": [None, 5, 10, 20],
            "model__min_samples_split": [2, 5, 10]
        }
    },
    "SVM": {
        "pipeline": Pipeline([
            ("pca", PCA(random_state=42)),
            ("model", SVC(class_weight="balanced", probability=True, random_state=42))
        ]),
        "params": {
            "pca__n_components": [0.80, 0.85, 0.90, 0.95, 0.99],
            "model__C": [0.1, 1, 10],
            "model__kernel": ["linear", "rbf"],
            "model__gamma": ["scale", "auto"]
        }
    },
    "KNN": {
        "pipeline": Pipeline([
            ("pca", PCA(random_state=42)),
            ("model", KNeighborsClassifier())
        ]),
        "params": {
            "pca__n_components": [0.80, 0.85, 0.90, 0.95, 0.99],
            "model__n_neighbors": [3, 5, 7, 9],
            "model__weights": ["uniform", "distance"],
            "model__metric": ["euclidean", "manhattan"]
        }
    },
    "Gradient Boosting": {
        "pipeline": Pipeline([
            ("pca", PCA(random_state=42)),
            ("model", GradientBoostingClassifier(random_state=42))
        ]),
        "params": {
            "pca__n_components": [0.80, 0.85, 0.90, 0.95, 0.99],
            "model__n_estimators": [100, 200],
            "model__learning_rate": [0.01, 0.1, 0.2],
            "model__max_depth": [3, 5, 10]
        }
    }
}

# -----------------------
# Step 3: Optimize All Models with RandomizedSearchCV
# -----------------------
best_models = {}
results = []

for name, mp in param_grids.items():
    print(f"\nOptimizing {name} ...")
    rs = RandomizedSearchCV(
        mp["pipeline"],
        mp["params"],
        scoring="f1",
        cv=5,
        n_iter=10,
        n_jobs=-1,
        random_state=42
    )
    rs.fit(X_train, y_train)
    best_models[name] = rs.best_estimator_
    results.append([name, rs.best_score_, rs.best_params_])
    print(f"Best {name}: F1={rs.best_score_:.4f}, Params={rs.best_params_}")

df_results = pd.DataFrame(results, columns=["Model", "Best CV F1", "Best Params"])
df_results = df_results.sort_values(by="Best CV F1", ascending=False).reset_index(drop=True)
df_results.to_excel("sprint3/AI_Electra_OptimizationResults.xlsx", index=False)

# -----------------------
# Step 4: Select Top 3 Models
# -----------------------
top3 = df_results.head(3)
best_model_name = top3.iloc[0]["Model"]
best_model = best_models[best_model_name]
print(f"\nBest Model Selected: {best_model_name}")

# -----------------------
# Step 5: Generate Predictions, Metrics, and Plots
# -----------------------
final_preds = pd.DataFrame()
metrics_summary = []
roc_data = {}

for _, row in top3.iterrows():
    name = row["Model"]
    model = best_models[name]
    
    # Fit model (PCA + classifier)
    model.fit(X_train, y_train)
    
    # Transform test set via PCA
    X_test_transformed = model.named_steps["pca"].transform(X_test)
    
    y_pred = model.named_steps["model"].predict(X_test_transformed)
    y_prob = model.named_steps["model"].predict_proba(X_test_transformed)[:,1] if hasattr(model.named_steps["model"], "predict_proba") else None
    final_preds[f"Predicted_{name.replace(' ','')}"] = y_pred
    
    # Metrics
    acc = accuracy_score(y_test, y_pred)
    prec = precision_score(y_test, y_pred)
    rec = recall_score(y_test, y_pred)
    f1 = f1_score(y_test, y_pred)
    tn, fp, fn, tp = confusion_matrix(y_test, y_pred).ravel()
    metrics_summary.append([name, acc, prec, rec, f1, tn, fp, fn, tp])
    
    # Confusion Matrix Plot
    cm = confusion_matrix(y_test, y_pred)
    disp = ConfusionMatrixDisplay(confusion_matrix=cm, display_labels=[0,1])
    disp.plot(cmap="Blues")
    plt.title(f"Confusion Matrix - {name}")
    cm_path = f"sprint3/Week5_Plots/ConfMatrix_{name.replace(' ','')}.png"
    plt.savefig(cm_path)
    plt.close()
    
    # ROC Curve Plot
    if y_prob is not None:
        fpr, tpr, _ = roc_curve(y_test, y_prob)
        roc_auc = auc(fpr, tpr)
        roc_data[name] = (fpr, tpr, roc_auc)
        plt.figure()
        plt.plot(fpr, tpr, label=f"{name} (AUC={roc_auc:.2f})")
        plt.plot([0,1],[0,1],"k--")
        plt.xlabel("False Positive Rate"); plt.ylabel("True Positive Rate")
        plt.title(f"ROC - {name}"); plt.legend()
        roc_path = f"sprint3/Week5_Plots/ROC_{name.replace(' ','')}.png"
        plt.savefig(roc_path)
        plt.close()

# Combined ROC for top 3
plt.figure(figsize=(8,6))
for name, (fpr, tpr, roc_auc) in roc_data.items():
    plt.plot(fpr, tpr, label=f"{name} (AUC={roc_auc:.2f})")
plt.plot([0,1],[0,1],"k--")
plt.xlabel("False Positive Rate"); plt.ylabel("True Positive Rate")
plt.title("Combined ROC - Top 3 Models")
plt.legend(loc="lower right")
combined_roc_path = "sprint3/Week5_Plots/ROC_Combined_Top3.png"
plt.savefig(combined_roc_path)
plt.close()

# -----------------------
# Step 6: Save Predictions and Metrics
# -----------------------
final_preds["Actual_Cancer"] = y_test.values
final_preds.to_excel("sprint3/AI_Electra_FinalPredictions.xlsx", index=False)

df_metrics = pd.DataFrame(metrics_summary, columns=["Model","Accuracy","Precision","Recall","F1","TN","FP","FN","TP"])
df_metrics.to_excel("sprint3/AI_Electra_FinalMetrics.xlsx", index=False)

# -----------------------
# Step 7: Create Deliverable Excel with Plots and Summary
# -----------------------
wb = Workbook()

# Sheet 1: Predictions
ws_preds = wb.active
ws_preds.title = "Final Predictions"
for r in dataframe_to_rows(final_preds, index=False, header=True):
    ws_preds.append(r)

# Sheet 2: Plots
ws_plots = wb.create_sheet("Plots")
row_pos = 2
for name in top3["Model"]:
    cm_img = XLImage(f"sprint3/Week5_Plots/ConfMatrix_{name.replace(' ','')}.png")
    roc_img = XLImage(f"sprint3/Week5_Plots/ROC_{name.replace(' ','')}.png")
    ws_plots[f"A{row_pos}"] = f"{name} Confusion & ROC"
    ws_plots.add_image(cm_img, f"A{row_pos+1}")
    ws_plots.add_image(roc_img, f"H{row_pos+1}")
    row_pos += 22

combined_img = XLImage(combined_roc_path)
ws_plots[f"A{row_pos}"] = "Combined ROC - Top 3"
ws_plots.add_image(combined_img, f"A{row_pos+1}")

# Sheet 3: Metrics Summary
ws_metrics = wb.create_sheet("Metrics Summary")
for r in dataframe_to_rows(df_metrics, index=False, header=True):
    ws_metrics.append(r)
for i, row in enumerate(df_metrics.itertuples(), start=2):
    if row.Model == best_model_name:
        ws_metrics[f"A{i}"].font = ws_metrics[f"A{i}"].font.copy(bold=True)

# Sheet 4: Summary
ws_summary = wb.create_sheet("Summary")
ws_summary["A1"] = "Week 5 Deliverable Summary - Cancer Detection"
ws_summary["A3"] = f"Total Records: {len(X_train)+len(X_test)}"
ws_summary["A4"] = f"Train Size: {len(X_train)}, Test Size: {len(X_test)}"
ws_summary["A5"] = "Class Distribution (0=No Cancer, 1=Cancer):"
ws_summary["A6"] = str(y_train.value_counts().to_dict() | y_test.value_counts().to_dict())
ws_summary["A8"] = "Dropped Columns: Patient_ID, First_Name, Last_Name, Doctor_Assigned, Admission_Date, Discharge_Date, Hospital_Department, Insurance_Provider, Tumor_Marker_Level"
ws_summary["A10"] = "Engineered Features: Systolic_BP, Diastolic_BP, BMI_Category, Hypertension_Flag"
ws_summary["A12"] = "Top 3 Models (by CV F1):"
for i, row in top3.iterrows():
    ws_summary[f"A{13+i}"] = f"{row['Model']} - CV F1: {row['Best CV F1']:.4f}"
ws_summary[f"A{16}"] = f"Best Model Selected: {best_model_name}"
ws_summary[f"A{17}"] = "Reason: Highest F1 score, balancing precision and recall for reliable cancer detection"
ws_summary[f"A{19}"] = "Notes: Tumor Marker excluded to avoid data leakage; Cancer detection task → F1 prioritized to minimize false negatives."

# Save Excel
wb.save("sprint3/AI_Electra_Week5_Deliverable.xlsx")
print("Week 5 deliverable Excel created: AI_Electra_Week5_Deliverable.xlsx")
